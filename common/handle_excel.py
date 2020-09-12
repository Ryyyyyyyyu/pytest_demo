# -*- coding: UTF-8 -*-
# @Time     :2020/9/9 16:49
# @Author   :raoyu

import openpyxl
import os
from common.handle_path import data_path


class Handle_Excel(object):
    """处理excel表格数据"""

    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def read_data(self):
        """读取excel文件数据"""
        wb = openpyxl.load_workbook(filename=self.filename)
        sh = wb[self.sheetname]
        rows_data = list(sh.rows)
        cases = []
        titles = []
        for title in rows_data[0]:
            titles.append(title.value)

        for datas in rows_data[1:]:
            values = []
            for itme in datas:
                values.append(itme.value)
            case = dict(zip(titles, values))
            cases.append(case)

        return cases

    def write_data(self, row, column, value):
        wb = openpyxl.load_workbook(filename=self.filename)
        sh = wb[self.sheetname]
        sh.cell(row=row, column=column, value=value)
        wb.save(self.filename)


if __name__ == '__main__':
    filename = os.path.join(data_path, 'apicases.xlsx')
    excel = Handle_Excel(filename, 'login')
    cases_data = excel.read_data()
    print(cases_data)
